"""
Tests P1 — Academic Structure & Import
"""
import pytest
from io import BytesIO
from rest_framework import status
from academic.models import Faculte, Departement, Filiere, Niveau, Salle
from conftest import auth_client


# ═══════════════════════════════════════════════════════
#  CRUD : Structure academique
# ═══════════════════════════════════════════════════════

@pytest.mark.django_db
class TestFaculte:

    def test_create_faculte(self, api, admin_user):
        c = auth_client(api, admin_user)
        res = c.post('/api/academic/facultes/', {'nom_faculte': 'Faculte de Droit'})
        assert res.status_code == status.HTTP_201_CREATED

    def test_list_facultes_public(self, api, structure):
        res = api.get('/api/academic/facultes/')
        assert res.status_code == status.HTTP_200_OK
        assert len(res.data) >= 1

    def test_enseignant_cannot_create_faculte(self, api, enseignant_user):
        c = auth_client(api, enseignant_user)
        res = c.post('/api/academic/facultes/', {'nom_faculte': 'Hack'})
        assert res.status_code == status.HTTP_403_FORBIDDEN


@pytest.mark.django_db
class TestDepartement:

    def test_create_departement(self, api, admin_user, structure):
        c = auth_client(api, admin_user)
        res = c.post('/api/academic/departements/', {
            'nom_departement': 'Mathematiques',
            'faculte': structure['faculte'].id,
        })
        assert res.status_code == status.HTTP_201_CREATED

    def test_assign_chef(self, api, admin_user, structure, chef_user):
        c = auth_client(api, admin_user)
        res = c.patch(f'/api/academic/departements/{structure["departement"].id}/', {
            'chef_departement': chef_user.id,
        }, format='json')
        assert res.status_code == status.HTTP_200_OK


@pytest.mark.django_db
class TestFiliere:

    def test_create_filiere(self, api, admin_user, structure):
        c = auth_client(api, admin_user)
        res = c.post('/api/academic/filieres/', {
            'nom_filiere': 'ICT',
            'departement': structure['departement'].id,
        })
        assert res.status_code == status.HTTP_201_CREATED

    def test_list_filieres_public(self, api, structure):
        res = api.get('/api/academic/filieres/')
        assert res.status_code == status.HTTP_200_OK


@pytest.mark.django_db
class TestNiveau:

    def test_create_niveau(self, api, admin_user, structure):
        c = auth_client(api, admin_user)
        # Create a new filiere first to avoid unique constraint
        fil = Filiere.objects.create(nom_filiere='ICT', departement=structure['departement'])
        res = c.post('/api/academic/niveaux/', {
            'nom_niveau': 'L1',
            'filiere': fil.id,
        })
        assert res.status_code == status.HTTP_201_CREATED

    def test_filter_niveaux_by_departement(self, api, admin_user, structure):
        c = auth_client(api, admin_user)
        res = c.get(f'/api/academic/niveaux/?departement={structure["departement"].id}')
        assert res.status_code == status.HTTP_200_OK
        assert len(res.data) >= 1


# ═══════════════════════════════════════════════════════
#  SALLES : CRUD + Import
# ═══════════════════════════════════════════════════════

@pytest.mark.django_db
class TestSalles:

    def test_create_salle(self, api, admin_user):
        c = auth_client(api, admin_user)
        res = c.post('/api/academic/salles/', {'nom_salle': 'PPE112'})
        assert res.status_code == status.HTTP_201_CREATED

    def test_duplicate_salle(self, api, admin_user, salle):
        c = auth_client(api, admin_user)
        res = c.post('/api/academic/salles/', {'nom_salle': 'A500'})
        assert res.status_code == status.HTTP_400_BAD_REQUEST

    def test_list_salles_public(self, api, salle):
        res = api.get('/api/academic/salles/')
        assert res.status_code == status.HTTP_200_OK

    def test_import_salles_excel(self, api, admin_user):
        import openpyxl
        c = auth_client(api, admin_user)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'nom_salle'
        ws['A2'] = 'R101'
        ws['A3'] = 'R108'
        ws['A4'] = 'S008'
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        buf.name = 'salles.xlsx'

        res = c.post('/api/academic/salles/import/', {'file': buf}, format='multipart')
        assert res.status_code == status.HTTP_200_OK
        assert res.data['created'] == 3
        assert Salle.objects.filter(nom_salle='R101').exists()

    def test_import_salles_skip_existing(self, api, admin_user, salle):
        import openpyxl
        c = auth_client(api, admin_user)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'nom_salle'
        ws['A2'] = 'A500'  # Already exists
        ws['A3'] = 'NewRoom'
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        buf.name = 'salles.xlsx'

        res = c.post('/api/academic/salles/import/', {'file': buf}, format='multipart')
        assert res.status_code == status.HTTP_200_OK
        assert res.data['created'] == 1
        assert res.data['skipped'] == 1

    def test_enseignant_cannot_import_salles(self, api, enseignant_user):
        c = auth_client(api, enseignant_user)
        res = c.post('/api/academic/salles/import/', {}, format='multipart')
        assert res.status_code == status.HTTP_403_FORBIDDEN


# ═══════════════════════════════════════════════════════
#  ANNEE ACADEMIQUE
# ═══════════════════════════════════════════════════════

@pytest.mark.django_db
class TestAnneeAcademique:

    def test_list_annees(self, api, admin_user, annee_active):
        c = auth_client(api, admin_user)
        res = c.get('/api/academic/annees-academiques/')
        assert res.status_code == status.HTTP_200_OK
        assert len(res.data) >= 1

    def test_annee_has_semestres(self, api, admin_user, annee_active):
        c = auth_client(api, admin_user)
        res = c.get(f'/api/academic/annees-academiques/{annee_active["annee"].id}/')
        assert res.status_code == status.HTTP_200_OK
        assert len(res.data['semestres']) == 2
