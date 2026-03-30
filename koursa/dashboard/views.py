from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework import status
from users.permissions import IsHoDOrSuperAdmin, IsSuperAdmin
from teaching.models import FicheSuivi, StatutFiche, UniteEnseignement
from academic.models import Departement, Filiere, Niveau, Semestre
from users.models import Utilisateur, Role
from django.db import models
from django.db.models import Sum, Count, Q
from django.core.cache import cache
from datetime import datetime, timedelta, date
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# === Helpers Excel ===

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="001EA6", end_color="001EA6", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)
CELL_ALIGNMENT = Alignment(vertical='center', wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)
TOTAL_FONT = Font(bold=True, size=11)
TOTAL_FILL = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")


def style_header_row(ws, row_num, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def style_data_cell(cell):
    cell.alignment = CELL_ALIGNMENT
    cell.border = THIN_BORDER


def get_departement(user):
    try:
        return Departement.objects.get(chef_departement=user)
    except Departement.DoesNotExist:
        return None


def resolve_departement(request):
    """Resolve the department for the current user.
    Super admins can specify ?departement=ID, otherwise all departments.
    HoDs always get their own department."""
    is_super = request.user.roles.filter(nom_role=Role.SUPER_ADMIN).exists()
    if is_super:
        dept_id = request.query_params.get('departement')
        if dept_id:
            try:
                return Departement.objects.get(id=int(dept_id))
            except (Departement.DoesNotExist, ValueError):
                return None
        return None  # None means all departments for super admin
    return get_departement(request.user)


def get_fiches_queryset(departement, date_debut=None, date_fin=None, filiere_id=None, niveau_id=None, semestre=None, annee_academique_id=None):
    base_filter = {'statut': StatutFiche.VALIDEE}
    if departement is not None:
        base_filter['ue__niveaux__filiere__departement'] = departement
    qs = FicheSuivi.objects.filter(
        **base_filter,
    ).select_related('ue', 'enseignant').distinct()
    if annee_academique_id:
        qs = qs.filter(semestre__annee_academique_id=annee_academique_id)
    if filiere_id:
        qs = qs.filter(ue__niveaux__filiere_id=filiere_id)
    if niveau_id:
        qs = qs.filter(ue__niveaux__id=niveau_id)
    if semestre:
        qs = qs.filter(ue__semestre=semestre)
    if date_debut:
        qs = qs.filter(date_cours__gte=date_debut)
    if date_fin:
        qs = qs.filter(date_cours__lte=date_fin)
    return qs.order_by('enseignant__last_name', 'enseignant__first_name', 'date_cours')


def parse_dates(request):
    date_debut = request.query_params.get('date_debut')
    date_fin = request.query_params.get('date_fin')
    try:
        if date_debut:
            date_debut = datetime.strptime(date_debut, '%Y-%m-%d').date()
        if date_fin:
            date_fin = datetime.strptime(date_fin, '%Y-%m-%d').date()
    except ValueError:
        return None, None, Response(
            {"detail": "Format de date invalide. Utilisez YYYY-MM-DD."},
            status=status.HTTP_400_BAD_REQUEST
        )
    return date_debut, date_fin, None


def parse_filters(request):
    filiere_id = request.query_params.get('filiere')
    niveau_id = request.query_params.get('niveau')
    semestre = request.query_params.get('semestre')
    try:
        filiere_id = int(filiere_id) if filiere_id else None
        niveau_id = int(niveau_id) if niveau_id else None
        semestre = int(semestre) if semestre else None
    except (ValueError, TypeError):
        filiere_id, niveau_id, semestre = None, None, None
    return filiere_id, niveau_id, semestre


def get_niveau_label(niveau_id, filiere_id, departement):
    """Build a label like INF_M1_S1 for filenames"""
    parts = []
    if filiere_id:
        try:
            filt = {'id': filiere_id}
            if departement is not None:
                filt['departement'] = departement
            filiere = Filiere.objects.get(**filt)
            parts.append(filiere.nom_filiere.replace(' ', '_'))
        except Filiere.DoesNotExist:
            pass
    if niveau_id:
        try:
            niveau = Niveau.objects.get(id=niveau_id)
            parts.append(niveau.nom_niveau.replace(' ', '_'))
        except Niveau.DoesNotExist:
            pass
    return '_'.join(parts) if parts else ''


def format_horaire(fiche):
    h_debut = fiche.heure_debut.strftime('%Hh%M') if fiche.heure_debut else ''
    h_fin = fiche.heure_fin.strftime('%Hh%M') if fiche.heure_fin else ''
    return f"{h_debut} - {h_fin}"


def format_contenu(fiche):
    prefix = fiche.type_seance
    titre = fiche.titre_chapitre or ''
    return f"{prefix} : {titre}"


def fiche_to_row(fiche, num):
    enseignant_nom = ''
    if fiche.enseignant:
        enseignant_nom = f"{fiche.enseignant.last_name} {fiche.enseignant.first_name}".strip()
    heures = round(fiche.duree.total_seconds() / 3600, 2) if fiche.duree else 0
    return [
        num,
        fiche.date_cours,
        format_horaire(fiche),
        enseignant_nom,
        fiche.ue.code_ue if fiche.ue else '',
        format_contenu(fiche),
        fiche.salle or '',
        heures,
    ]


BILAN_HEADERS = ["No", "Date", "Horaire", "Enseignant", "UE", "Contenu", "Salle", "Total horaire"]
BILAN_COL_WIDTHS = [6, 14, 18, 22, 12, 50, 16, 14]


def write_bilan_sheet(ws, fiches, title=None):
    if title:
        ws.title = title

    ws.append(BILAN_HEADERS)
    style_header_row(ws, 1, len(BILAN_HEADERS))

    total_heures = 0
    for i, fiche in enumerate(fiches, 1):
        row_data = fiche_to_row(fiche, i)
        ws.append(row_data)
        total_heures += row_data[7]
        for col in range(1, len(BILAN_HEADERS) + 1):
            style_data_cell(ws.cell(row=i + 1, column=col))
        # Format date
        ws.cell(row=i + 1, column=2).number_format = 'DD/MM/YYYY'

    # Total row
    total_row = ws.max_row + 1
    ws.cell(row=total_row, column=7, value="TOTAL")
    ws.cell(row=total_row, column=8, value=round(total_heures, 2))
    for col in range(1, len(BILAN_HEADERS) + 1):
        cell = ws.cell(row=total_row, column=col)
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
        cell.border = THIN_BORDER

    # Column widths
    for i, w in enumerate(BILAN_COL_WIDTHS, 1):
        ws.column_dimensions[chr(64 + i)].width = w


def excel_response(wb, filename):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# === Views ===

class DashboardRootView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        return Response({
            "message": "Bienvenue sur le Dashboard Koursa",
            "endpoints": {
                "stats": "/api/dashboard/stats/",
                "recapitulatif": "/api/dashboard/recapitulatif/",
                "export_bilan": "/api/dashboard/export-bilan/",
                "export_par_ue": "/api/dashboard/export-par-ue/",
                "export_par_enseignant": "/api/dashboard/export-par-enseignant/",
                "export_heures": "/api/dashboard/export-heures/",
            }
        })


class DashboardStatsView(APIView):
    permission_classes = [IsAuthenticated, IsHoDOrSuperAdmin]

    def get(self, request, *args, **kwargs):
        departement = resolve_departement(request)
        is_super = request.user.roles.filter(nom_role=Role.SUPER_ADMIN).exists()
        # HoD must have a department
        if not is_super and not departement:
            return Response(
                {"detail": "Vous n'etes assigne a aucun departement."},
                status=status.HTTP_404_NOT_FOUND
            )

        filiere_id, niveau_id, semestre = parse_filters(request)

        current_month = datetime.now().month
        current_year = datetime.now().year

        # Base queryset filtered by dept + niveau/filiere/semestre
        base_filter = {}
        if departement is not None:
            base_filter['ue__niveaux__filiere__departement'] = departement
        if filiere_id:
            base_filter['ue__niveaux__filiere_id'] = filiere_id
        if niveau_id:
            base_filter['ue__niveaux__id'] = niveau_id
        if semestre:
            base_filter['ue__semestre'] = semestre

        heures_validees_mois = FicheSuivi.objects.filter(
            **base_filter,
            statut=StatutFiche.VALIDEE,
            date_cours__year=current_year,
            date_cours__month=current_month
        ).distinct().aggregate(total_heures=Sum('duree'))['total_heures'] or timedelta(0)

        cutoff_date = datetime.now() - timedelta(days=2)
        fiches_en_retard = FicheSuivi.objects.filter(
            **base_filter,
            statut=StatutFiche.SOUMISE,
            date_soumission__lt=cutoff_date
        ).distinct().count()

        heures_par_ue = FicheSuivi.objects.filter(
            **base_filter,
            statut=StatutFiche.VALIDEE,
            date_cours__year=current_year,
            date_cours__month=current_month
        ).distinct().values('ue__code_ue', 'ue__libelle_ue').annotate(
            total_duree=Sum('duree')
        ).order_by('-total_duree')

        # Filieres et niveaux disponibles pour les selecteurs
        if departement is not None:
            filieres = Filiere.objects.filter(departement=departement).order_by('nom_filiere')
            niveaux = Niveau.objects.filter(filiere__departement=departement).select_related('filiere').order_by('filiere__nom_filiere', 'nom_niveau')
        else:
            filieres = Filiere.objects.all().order_by('nom_filiere')
            niveaux = Niveau.objects.all().select_related('filiere').order_by('filiere__nom_filiere', 'nom_niveau')

        data = {
            'heures_validees_ce_mois': round(heures_validees_mois.total_seconds() / 3600, 2),
            'fiches_en_retard_de_validation': fiches_en_retard,
            'repartition_heures_par_ue_ce_mois': [
                {
                    'code_ue': item['ue__code_ue'],
                    'libelle_ue': item['ue__libelle_ue'],
                    'heures_effectuees': round(item['total_duree'].total_seconds() / 3600, 2)
                } for item in heures_par_ue
            ],
            'filieres': [
                {'id': f.id, 'nom': f.nom_filiere} for f in filieres
            ],
            'niveaux': [
                {'id': n.id, 'nom': n.nom_niveau, 'filiere_id': n.filiere_id, 'filiere_nom': n.filiere.nom_filiere} for n in niveaux
            ],
        }

        return Response(data)


class RecapitulatifView(APIView):
    """Recapitulatif JSON des heures par UE et par enseignant"""
    permission_classes = [IsAuthenticated, IsHoDOrSuperAdmin]

    def get(self, request, *args, **kwargs):
        departement = resolve_departement(request)
        is_super = request.user.roles.filter(nom_role=Role.SUPER_ADMIN).exists()
        if not is_super and not departement:
            return Response({"detail": "Vous n'etes assigne a aucun departement."}, status=status.HTTP_404_NOT_FOUND)

        date_debut, date_fin, error = parse_dates(request)
        if error:
            return error

        filiere_id, niveau_id, semestre = parse_filters(request)
        fiches = get_fiches_queryset(departement, date_debut, date_fin, filiere_id, niveau_id, semestre)

        # Par UE
        par_ue = fiches.values('ue__code_ue', 'ue__libelle_ue').annotate(
            total_heures=Sum('duree'),
            nb_fiches=Count('id'),
        ).order_by('ue__code_ue')

        # Par enseignant
        par_enseignant = fiches.values(
            'enseignant__id', 'enseignant__first_name', 'enseignant__last_name'
        ).annotate(
            total_heures=Sum('duree'),
            nb_fiches=Count('id'),
        ).order_by('enseignant__last_name')

        total_heures_global = fiches.aggregate(t=Sum('duree'))['t'] or timedelta(0)

        # Filieres et niveaux disponibles
        if departement is not None:
            filieres = Filiere.objects.filter(departement=departement).order_by('nom_filiere')
            niveaux = Niveau.objects.filter(filiere__departement=departement).select_related('filiere').order_by('filiere__nom_filiere', 'nom_niveau')
        else:
            filieres = Filiere.objects.all().order_by('nom_filiere')
            niveaux = Niveau.objects.all().select_related('filiere').order_by('filiere__nom_filiere', 'nom_niveau')

        data = {
            'total_heures': round(total_heures_global.total_seconds() / 3600, 2),
            'total_fiches': fiches.count(),
            'par_ue': [
                {
                    'code_ue': item['ue__code_ue'],
                    'libelle_ue': item['ue__libelle_ue'],
                    'heures': round(item['total_heures'].total_seconds() / 3600, 2) if item['total_heures'] else 0,
                    'nb_fiches': item['nb_fiches'],
                } for item in par_ue
            ],
            'par_enseignant': [
                {
                    'id': item['enseignant__id'],
                    'nom': f"{item['enseignant__last_name']} {item['enseignant__first_name']}".strip(),
                    'heures': round(item['total_heures'].total_seconds() / 3600, 2) if item['total_heures'] else 0,
                    'nb_fiches': item['nb_fiches'],
                } for item in par_enseignant
            ],
            'filieres': [
                {'id': f.id, 'nom': f.nom_filiere} for f in filieres
            ],
            'niveaux': [
                {'id': n.id, 'nom': n.nom_niveau, 'filiere_id': n.filiere_id, 'filiere_nom': n.filiere.nom_filiere} for n in niveaux
            ],
        }

        return Response(data)


class ExportBilanView(APIView):
    """Export global au format bilan (identique au fichier de reference)"""
    permission_classes = [IsAuthenticated, IsHoDOrSuperAdmin]

    def get(self, request, *args, **kwargs):
        departement = resolve_departement(request)
        is_super = request.user.roles.filter(nom_role=Role.SUPER_ADMIN).exists()
        if not is_super and not departement:
            return Response({"detail": "Vous n'etes assigne a aucun departement."}, status=status.HTTP_404_NOT_FOUND)

        date_debut, date_fin, error = parse_dates(request)
        if error:
            return error

        filiere_id, niveau_id, semestre = parse_filters(request)
        fiches = get_fiches_queryset(departement, date_debut, date_fin, filiere_id, niveau_id, semestre)

        wb = Workbook()
        ws = wb.active

        # Build filename like: BILAN_DES_COURS_INF_M1_S1_2025-2026.xlsx
        name_parts = ["BILAN_DES_COURS"]
        level_label = get_niveau_label(niveau_id, filiere_id, departement)
        if level_label:
            name_parts.append(level_label)
        if semestre:
            name_parts.append(f"S{semestre}")
        if date_debut and date_fin:
            name_parts.append(f"{date_debut}_{date_fin}")

        sheet_title = ' '.join(["Bilan des cours"] + ([level_label] if level_label else []))
        write_bilan_sheet(ws, fiches, title=sheet_title[:31])

        filename = '_'.join(name_parts) + '.xlsx'
        return excel_response(wb, filename)


class ExportParUEView(APIView):
    """Export par UE - une UE specifique ou toutes les UEs (un onglet par UE)"""
    permission_classes = [IsAuthenticated, IsHoDOrSuperAdmin]

    def get(self, request, *args, **kwargs):
        departement = resolve_departement(request)
        is_super = request.user.roles.filter(nom_role=Role.SUPER_ADMIN).exists()
        if not is_super and not departement:
            return Response({"detail": "Vous n'etes assigne a aucun departement."}, status=status.HTTP_404_NOT_FOUND)

        date_debut, date_fin, error = parse_dates(request)
        if error:
            return error

        filiere_id, niveau_id, semestre = parse_filters(request)
        ue_id = request.query_params.get('ue')
        fiches = get_fiches_queryset(departement, date_debut, date_fin, filiere_id, niveau_id, semestre)

        wb = Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        if ue_id:
            # Export une seule UE
            fiches_ue = fiches.filter(ue_id=ue_id)
            if not fiches_ue.exists():
                return Response({"detail": "Aucune fiche pour cette UE."}, status=status.HTTP_404_NOT_FOUND)
            code_ue = fiches_ue.first().ue.code_ue
            ws = wb.create_sheet(title=code_ue[:31])
            write_bilan_sheet(ws, fiches_ue)
            filename = f"fiches_{code_ue}.xlsx"
        else:
            # Export toutes les UEs - un onglet par UE
            ue_codes = fiches.values_list('ue__code_ue', 'ue_id').distinct().order_by('ue__code_ue')
            if not ue_codes:
                return Response({"detail": "Aucune fiche trouvee."}, status=status.HTTP_404_NOT_FOUND)
            for code_ue, ue_id_val in ue_codes:
                fiches_ue = fiches.filter(ue_id=ue_id_val)
                ws = wb.create_sheet(title=code_ue[:31])
                write_bilan_sheet(ws, fiches_ue)

            # Onglet recapitulatif - single aggregated query
            ws_recap = wb.create_sheet(title="Recapitulatif", index=0)
            recap_headers = ["UE", "Nb seances", "Total heures"]
            ws_recap.append(recap_headers)
            style_header_row(ws_recap, 1, len(recap_headers))
            total_h = 0
            total_s = 0
            recap_data = fiches.values('ue__code_ue').annotate(
                nb=Count('id'), total_duree=Sum('duree')
            ).order_by('ue__code_ue')
            for item in recap_data:
                h = item['total_duree'] or timedelta(0)
                heures = round(h.total_seconds() / 3600, 2)
                total_h += heures
                total_s += item['nb']
                ws_recap.append([item['ue__code_ue'], item['nb'], heures])
                for col in range(1, 4):
                    style_data_cell(ws_recap.cell(row=ws_recap.max_row, column=col))
            # Total
            ws_recap.append(["TOTAL", total_s, round(total_h, 2)])
            for col in range(1, 4):
                cell = ws_recap.cell(row=ws_recap.max_row, column=col)
                cell.font = TOTAL_FONT
                cell.fill = TOTAL_FILL
                cell.border = THIN_BORDER
            ws_recap.column_dimensions['A'].width = 14
            ws_recap.column_dimensions['B'].width = 14
            ws_recap.column_dimensions['C'].width = 14

            dept_label = departement.nom_departement.lower().replace(' ', '_') if departement else 'tous'
            filename = f"fiches_par_ue_{dept_label}.xlsx"

        return excel_response(wb, filename)


class ExportParEnseignantView(APIView):
    """Export par enseignant - un enseignant specifique ou tous (un onglet par enseignant)"""
    permission_classes = [IsAuthenticated, IsHoDOrSuperAdmin]

    def get(self, request, *args, **kwargs):
        departement = resolve_departement(request)
        is_super = request.user.roles.filter(nom_role=Role.SUPER_ADMIN).exists()
        if not is_super and not departement:
            return Response({"detail": "Vous n'etes assigne a aucun departement."}, status=status.HTTP_404_NOT_FOUND)

        date_debut, date_fin, error = parse_dates(request)
        if error:
            return error

        filiere_id, niveau_id, semestre = parse_filters(request)
        enseignant_id = request.query_params.get('enseignant')
        fiches = get_fiches_queryset(departement, date_debut, date_fin, filiere_id, niveau_id, semestre)

        wb = Workbook()
        wb.remove(wb.active)

        if enseignant_id:
            fiches_ens = fiches.filter(enseignant_id=enseignant_id)
            if not fiches_ens.exists():
                return Response({"detail": "Aucune fiche pour cet enseignant."}, status=status.HTTP_404_NOT_FOUND)
            ens = fiches_ens.first().enseignant
            nom = f"{ens.last_name} {ens.first_name}".strip()
            ws = wb.create_sheet(title=nom[:31])
            write_bilan_sheet(ws, fiches_ens)
            filename = f"fiches_{nom.replace(' ', '_')}.xlsx"
        else:
            enseignants = fiches.values_list(
                'enseignant__id', 'enseignant__last_name', 'enseignant__first_name'
            ).distinct().order_by('enseignant__last_name')

            if not enseignants:
                return Response({"detail": "Aucune fiche trouvee."}, status=status.HTTP_404_NOT_FOUND)

            for ens_id, last, first in enseignants:
                fiches_ens = fiches.filter(enseignant_id=ens_id)
                nom = f"{last} {first}".strip()
                ws = wb.create_sheet(title=nom[:31])
                write_bilan_sheet(ws, fiches_ens)

            # Onglet recapitulatif - single aggregated query
            ws_recap = wb.create_sheet(title="Recapitulatif", index=0)
            recap_headers = ["Enseignant", "Nb seances", "Total heures"]
            ws_recap.append(recap_headers)
            style_header_row(ws_recap, 1, len(recap_headers))
            total_h = 0
            total_s = 0
            recap_data = fiches.values(
                'enseignant__last_name', 'enseignant__first_name'
            ).annotate(
                nb=Count('id'), total_duree=Sum('duree')
            ).order_by('enseignant__last_name')
            for item in recap_data:
                h = item['total_duree'] or timedelta(0)
                heures = round(h.total_seconds() / 3600, 2)
                total_h += heures
                total_s += item['nb']
                nom = f"{item['enseignant__last_name']} {item['enseignant__first_name']}".strip()
                ws_recap.append([nom, item['nb'], heures])
                for col in range(1, 4):
                    style_data_cell(ws_recap.cell(row=ws_recap.max_row, column=col))
            ws_recap.append(["TOTAL", total_s, round(total_h, 2)])
            for col in range(1, 4):
                cell = ws_recap.cell(row=ws_recap.max_row, column=col)
                cell.font = TOTAL_FONT
                cell.fill = TOTAL_FILL
                cell.border = THIN_BORDER
            ws_recap.column_dimensions['A'].width = 25
            ws_recap.column_dimensions['B'].width = 14
            ws_recap.column_dimensions['C'].width = 14

            dept_label = departement.nom_departement.lower().replace(' ', '_') if departement else 'tous'
            filename = f"fiches_par_enseignant_{dept_label}.xlsx"

        return excel_response(wb, filename)


class ExportHeuresView(APIView):
    """Export ancien (compatibilite) - heures par enseignant pour un mois donne"""
    permission_classes = [IsAuthenticated, IsHoDOrSuperAdmin]

    def get(self, request, *args, **kwargs):
        departement = resolve_departement(request)
        is_super = request.user.roles.filter(nom_role=Role.SUPER_ADMIN).exists()
        if not is_super and not departement:
            return Response({"detail": "Vous n'etes assigne a aucun departement."}, status=status.HTTP_404_NOT_FOUND)

        try:
            year = int(request.query_params.get('annee', datetime.now().year))
            month = int(request.query_params.get('mois', datetime.now().month))
        except (ValueError, TypeError):
            return Response({"detail": "Parametres annee/mois invalides."}, status=status.HTTP_400_BAD_REQUEST)

        base_filter = {
            'statut': StatutFiche.VALIDEE,
            'date_cours__year': year,
            'date_cours__month': month,
        }
        if departement is not None:
            base_filter['ue__niveaux__filiere__departement'] = departement

        heures_par_enseignant = FicheSuivi.objects.filter(
            **base_filter,
        ).values(
            'enseignant__id',
            'enseignant__first_name',
            'enseignant__last_name'
        ).annotate(
            total_duree=Sum('duree')
        ).order_by('enseignant__last_name')

        wb = Workbook()
        ws = wb.active
        ws.title = f"Heures {month:02d}-{year}"

        headers = ["Nom", "Prenom", "Heures Effectuees"]
        ws.append(headers)
        style_header_row(ws, 1, len(headers))

        for item in heures_par_enseignant:
            heures = round(item['total_duree'].total_seconds() / 3600, 2) if item['total_duree'] else 0
            ws.append([
                item['enseignant__last_name'],
                item['enseignant__first_name'],
                heures
            ])

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 18

        dept_label = departement.nom_departement.lower() if departement else 'tous'
        filename = f"export_heures_{dept_label}_{month:02d}_{year}.xlsx"
        return excel_response(wb, filename)


class ExportMonRapportView(APIView):
    """Export personnel pour l'enseignant connecte - ses fiches validees"""
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        user = request.user
        if not user.roles.filter(nom_role__in=[Role.ENSEIGNANT, Role.CHEF_DEPARTEMENT]).exists():
            return Response(
                {"detail": "Acces reserve aux enseignants."},
                status=status.HTTP_403_FORBIDDEN,
            )

        date_debut, date_fin, error = parse_dates(request)
        if error:
            return error

        filiere_id, niveau_id, semestre = parse_filters(request)
        annee_academique_id = request.query_params.get('annee_academique')

        # Fiches validees de cet enseignant uniquement
        qs = FicheSuivi.objects.filter(
            enseignant=user,
            statut=StatutFiche.VALIDEE,
        ).select_related('ue', 'enseignant').distinct()

        if annee_academique_id:
            qs = qs.filter(semestre__annee_academique_id=annee_academique_id)
        if filiere_id:
            qs = qs.filter(ue__niveaux__filiere_id=filiere_id)
        if niveau_id:
            qs = qs.filter(ue__niveaux__id=niveau_id)
        if semestre:
            qs = qs.filter(ue__semestre=semestre)
        if date_debut:
            qs = qs.filter(date_cours__gte=date_debut)
        if date_fin:
            qs = qs.filter(date_cours__lte=date_fin)

        qs = qs.order_by('date_cours')

        if not qs.exists():
            return Response(
                {"detail": "Aucune fiche validee trouvee."},
                status=status.HTTP_404_NOT_FOUND,
            )

        wb = Workbook()
        ws = wb.active

        nom_enseignant = f"{user.last_name} {user.first_name}".strip()
        name_parts = ["MON_RAPPORT", nom_enseignant.replace(' ', '_')]
        level_label = get_niveau_label(niveau_id, filiere_id, None)
        if level_label:
            name_parts.append(level_label)
        if semestre:
            name_parts.append(f"S{semestre}")

        sheet_title = f"Rapport {nom_enseignant}"
        write_bilan_sheet(ws, qs, title=sheet_title[:31])

        filename = '_'.join(name_parts) + '.xlsx'
        return excel_response(wb, filename)


class AdminOverviewView(APIView):
    """Vue d'ensemble pour le super admin : stats par departement"""
    permission_classes = [IsAuthenticated, IsSuperAdmin]

    def get(self, request, *args, **kwargs):
        # Cache for 2 minutes - heavy query
        cache_key = 'admin_overview_stats'
        cached = cache.get(cache_key)
        if cached:
            return Response(cached)

        departements = Departement.objects.select_related('faculte', 'chef_departement').all()

        # Single aggregated query instead of N+1
        from django.db.models import Case, When, IntegerField
        stats = FicheSuivi.objects.filter(
            ue__niveaux__filiere__departement__isnull=False,
        ).values(
            'ue__niveaux__filiere__departement'
        ).annotate(
            total=Count('id', distinct=True),
            soumises=Count('id', distinct=True, filter=models.Q(statut=StatutFiche.SOUMISE)),
            validees=Count('id', distinct=True, filter=models.Q(statut=StatutFiche.VALIDEE)),
            refusees=Count('id', distinct=True, filter=models.Q(statut=StatutFiche.REFUSEE)),
        )
        stats_map = {s['ue__niveaux__filiere__departement']: s for s in stats}

        result = []
        for dept in departements:
            s = stats_map.get(dept.id, {})
            chef = dept.chef_departement
            result.append({
                'id': dept.id,
                'nom': dept.nom_departement,
                'faculte': dept.faculte.nom_faculte if dept.faculte else None,
                'chef': f"{chef.last_name} {chef.first_name}".strip() if chef else None,
                'total': s.get('total', 0),
                'soumises': s.get('soumises', 0),
                'validees': s.get('validees', 0),
                'refusees': s.get('refusees', 0),
            })

        # Totaux globaux - single query with conditional aggregation
        totaux = FicheSuivi.objects.aggregate(
            total=Count('id'),
            soumises=Count('id', filter=models.Q(statut=StatutFiche.SOUMISE)),
            validees=Count('id', filter=models.Q(statut=StatutFiche.VALIDEE)),
            refusees=Count('id', filter=models.Q(statut=StatutFiche.REFUSEE)),
        )

        response_data = {
            'departements': result,
            'totaux': totaux,
        }
        cache.set(cache_key, response_data, 120)  # 2 minutes
        return Response(response_data)


def get_monday(d=None):
    """Retourne le lundi de la semaine pour une date donnee."""
    if d is None:
        d = date.today()
    return d - timedelta(days=d.weekday())


class WeeklyTrackingView(APIView):
    """Suivi hebdomadaire pour le chef de departement."""
    permission_classes = [IsAuthenticated, IsHoDOrSuperAdmin]

    def get(self, request):
        # Parse semaine param
        semaine_str = request.query_params.get('semaine')
        if semaine_str:
            try:
                monday = datetime.strptime(semaine_str, '%Y-%m-%d').date()
            except ValueError:
                return Response(
                    {'detail': 'Format de date invalide. Utilisez YYYY-MM-DD.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        else:
            monday = get_monday()

        sunday = monday + timedelta(days=6)

        departement = resolve_departement(request)
        is_super = request.user.roles.filter(nom_role=Role.SUPER_ADMIN).exists()
        if not is_super and not departement:
            return Response(
                {'detail': 'Vous n\'etes assigne a aucun departement.'},
                status=status.HTTP_404_NOT_FOUND,
            )

        # Trouver le semestre actif
        semestre_actif = Semestre.objects.filter(est_actif=True).first()
        if not semestre_actif:
            return Response(
                {'detail': 'Aucun semestre actif.'},
                status=status.HTTP_404_NOT_FOUND,
            )

        # Construire la structure par filiere/niveau/UE
        if departement is not None:
            filieres = Filiere.objects.filter(departement=departement).order_by('nom_filiere')
        else:
            filieres = Filiere.objects.all().order_by('nom_filiere')

        stats_total = 0
        stats_validees = 0
        stats_soumises = 0
        stats_manquantes = 0

        filieres_data = []
        for filiere in filieres:
            niveaux = Niveau.objects.filter(filiere=filiere).order_by('nom_niveau')
            niveaux_data = []
            for niveau in niveaux:
                ues = UniteEnseignement.objects.filter(
                    niveaux=niveau,
                    semestre_obj=semestre_actif,
                ).distinct().order_by('code_ue')

                ues_data = []
                for ue in ues:
                    fiches = FicheSuivi.objects.filter(
                        ue=ue,
                        date_cours__gte=monday,
                        date_cours__lte=sunday,
                    ).order_by('date_cours')

                    has_validee = fiches.filter(statut=StatutFiche.VALIDEE).exists()
                    has_soumise = fiches.filter(statut=StatutFiche.SOUMISE).exists()

                    if has_validee:
                        statut = 'validee'
                        stats_validees += 1
                    elif has_soumise:
                        statut = 'soumise'
                        stats_soumises += 1
                    else:
                        statut = 'aucune_fiche'
                        stats_manquantes += 1
                    stats_total += 1

                    enseignants = ue.enseignants.all()
                    enseignants_data = [
                        {
                            'id': e.id,
                            'nom': f"{e.last_name} {e.first_name}".strip(),
                            'email': e.email,
                        }
                        for e in enseignants
                    ]

                    # Trouver le delegue du niveau
                    delegue = Utilisateur.objects.filter(niveau_represente=niveau).first()
                    delegue_data = None
                    if delegue:
                        delegue_data = {
                            'id': delegue.id,
                            'nom': f"{delegue.last_name} {delegue.first_name}".strip(),
                        }

                    fiches_data = [
                        {
                            'id': f.id,
                            'statut': f.statut,
                            'date_cours': str(f.date_cours),
                        }
                        for f in fiches
                    ]

                    ues_data.append({
                        'id': ue.id,
                        'code_ue': ue.code_ue,
                        'libelle_ue': ue.libelle_ue,
                        'enseignants': enseignants_data,
                        'delegue': delegue_data,
                        'statut': statut,
                        'fiches': fiches_data,
                    })

                if ues_data:
                    niveaux_data.append({
                        'id': niveau.id,
                        'nom': niveau.nom_niveau,
                        'ues': ues_data,
                    })

            if niveaux_data:
                filieres_data.append({
                    'id': filiere.id,
                    'nom': filiere.nom_filiere,
                    'niveaux': niveaux_data,
                })

        return Response({
            'semaine': str(monday),
            'departement': departement.nom_departement if departement else 'Tous',
            'filieres': filieres_data,
            'stats': {
                'total_ues': stats_total,
                'validees': stats_validees,
                'soumises': stats_soumises,
                'manquantes': stats_manquantes,
            },
        })


class EnseignantWeeklyTrackingView(APIView):
    """Suivi hebdomadaire pour l'enseignant connecte."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        semaine_str = request.query_params.get('semaine')
        if semaine_str:
            try:
                monday = datetime.strptime(semaine_str, '%Y-%m-%d').date()
            except ValueError:
                return Response(
                    {'detail': 'Format de date invalide. Utilisez YYYY-MM-DD.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        else:
            monday = get_monday()

        sunday = monday + timedelta(days=6)

        # Semestre actif
        semestre_actif = Semestre.objects.filter(est_actif=True).first()
        if not semestre_actif:
            return Response(
                {'detail': 'Aucun semestre actif.'},
                status=status.HTTP_404_NOT_FOUND,
            )

        # UEs de l'enseignant pour le semestre actif
        ues = UniteEnseignement.objects.filter(
            enseignants=request.user,
            semestre_obj=semestre_actif,
        ).distinct().order_by('code_ue')

        stats_total = 0
        stats_validees = 0
        stats_soumises = 0
        stats_manquantes = 0

        ues_data = []
        for ue in ues:
            fiches = FicheSuivi.objects.filter(
                ue=ue,
                date_cours__gte=monday,
                date_cours__lte=sunday,
            ).order_by('date_cours')

            has_validee = fiches.filter(statut=StatutFiche.VALIDEE).exists()
            has_soumise = fiches.filter(statut=StatutFiche.SOUMISE).exists()

            if has_validee:
                statut = 'validee'
                stats_validees += 1
            elif has_soumise:
                statut = 'soumise'
                stats_soumises += 1
            else:
                statut = 'aucune_fiche'
                stats_manquantes += 1
            stats_total += 1

            # Niveaux de cette UE
            niveaux = ue.niveaux.all().select_related('filiere')
            niveaux_labels = [
                f"{n.nom_niveau} {n.filiere.nom_filiere}" for n in niveaux
            ]

            # Delegues des niveaux lies a cette UE
            delegues = Utilisateur.objects.filter(
                niveau_represente__in=ue.niveaux.all()
            ).distinct()
            delegues_data = [
                {
                    'id': d.id,
                    'nom': f"{d.last_name} {d.first_name}".strip(),
                    'niveau': f"{d.niveau_represente.nom_niveau} {d.niveau_represente.filiere.nom_filiere}" if d.niveau_represente else '',
                }
                for d in delegues
            ]

            fiches_data = [
                {
                    'id': f.id,
                    'statut': f.statut,
                    'date_cours': str(f.date_cours),
                }
                for f in fiches
            ]

            ues_data.append({
                'id': ue.id,
                'code_ue': ue.code_ue,
                'libelle_ue': ue.libelle_ue,
                'niveaux': niveaux_labels,
                'statut': statut,
                'fiches': fiches_data,
                'delegues': delegues_data,
            })

        return Response({
            'semaine': str(monday),
            'ues': ues_data,
            'stats': {
                'total': stats_total,
                'validees': stats_validees,
                'soumises': stats_soumises,
                'manquantes': stats_manquantes,
            },
        })
